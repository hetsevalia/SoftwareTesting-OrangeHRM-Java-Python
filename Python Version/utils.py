import time
import pyautogui
from openpyxl import load_workbook

def get_cell_value(cell):
    return str(cell.value) if cell else ""

def read_excel_data(filepath):
    wb = load_workbook(filepath)
    data = []

    row = []
    sheet_cells = [
        ("Login", 1, [1, 2]),
        ("AddEmployee", 1, list(range(1, 7))),
        ("PersonalDetails", 1, list(range(1, 10))),
        ("SystemUser", 1, list(range(1, 6))),
        ("AssignLeave", 1, list(range(1, 6))),
        ("EmployeePunch", 1, [1, 2]),
        ("EmployeeClaim", 1, [1, 2, 3]),
        ("ExpenseInfo", 1, [1, 2, 3, 4]),
        ("BuzzFeed", 1, [1])
    ]

    for sheet, rownum, columns in sheet_cells:
        sheet_obj = wb[sheet]
        for col in columns:
            row.append(get_cell_value(sheet_obj.cell(row=rownum + 1, column=col)))

    wb.close()
    data.append(tuple(row))
    return data

def upload_file(path):
    time.sleep(1)
    pyautogui.write(path)
    pyautogui.press("enter")
    time.sleep(1)